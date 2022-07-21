# 導入系統模組
import sys
import re
import time
import datetime
import itertools
# 導入Excel模組
from openpyxl import load_workbook
# 導入網路爬蟲模組
import requests
from bs4 import BeautifulSoup
# 導入詞語分析模組
from collections import Counter
import jieba
# 導入資料庫模組
import sqlite3


# 全域變數設定


# 中華大學網站 爬蟲副程式
def cra():
    # ==========================================================================
    # 起始 - 執行時間
    start = datetime.datetime.now()
    # ==========================================================================
    # 爬蟲套件 模擬使用者瀏覽措施 自訂HTTP Headers
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                             'Chrome/74.0.3729.169 Safari/537.36'}
    # ==========================================================================
    # Excel套件 導入試算表
    wb = load_workbook('resource/cra_data.xlsx')
    # ==========================================================================
    # 結巴套件 中文字典、自訂字典、停用詞字典
    jieba.set_dictionary('resource/ji_dict_big.txt')
    jieba.load_userdict('resource/ji_dict.txt')
    with open('resource/ji_remove_dict.txt', encoding='utf-8') as f:
        stopWords = [line.strip() for line in f.readlines()]
    # ==========================================================================
    # 爬蟲套件 連結統整處
    URL_acad = 'http://www1.chu.edu.tw/p/412-1000-565.php?Lang=zh-tw'
    URL_offi = 'http://www1.chu.edu.tw/p/412-1000-495.php?Lang=zh-tw'
    # ==========================================================================
    # cra 區域變數
    URL_ac_of = ''
    tmp: str = ''
    tmp01: str = ''
    # =====================
    Ac_List = []  # 學術單位名稱
    Ac_WEB_List = []  # 學術單位網路位置
    Of_List = []  # 行政單位名稱
    Of_WEB_List = []  # 行政單位網路位置
    # =====================
    Ac_unity_text = []  # 單位用文字
    Ac_unity_Web = []  # 單位用網址
    ac_count = []
    Of_unity_text = []  # 單位用文字
    Of_unity_Web = []  # 單位用網址
    of_count = []
    # =====================
    tmp03 = ''
    tmp05 = []
    tmp06 = []
    tmp10 = []
    tmp11 = []
    tmp12 = ''
    tmp14 = ''
    # 學術總單位
    point_ac_all_text = []
    point_ac_all_freq = []
    ratio_ac_all_freq = []
    ac_all_sentence = []  # 學術總單位關鍵字句子
    # 學術分單位
    point_ac_dep_text = []
    point_ac_dep_freq = []
    ratio_ac_dep_freq = []
    ac_dep_sentence = []  # 學術分單位關鍵字句子
    # 行政總單位
    point_of_all_text = []
    point_of_all_freq = []
    ratio_of_all_freq = []
    of_all_sentence = []  # 行政總單位關鍵字句子
    # 行政分單位
    point_of_dep_text = []
    point_of_dep_freq = []
    ratio_of_dep_freq = []
    of_dep_sentence = []  # 行政分單位關鍵字句子
    # =====================
    cc = 0
    # ==========================================================================
    print('========校園官網_爬蟲程式　開始========')
    # =================學術單位=================
    Academic = requests.get(URL_acad)
    BS_2_3 = BeautifulSoup(Academic.text, 'html.parser')
    # 資訊電機學院
    Data_2_5_1 = BS_2_3.find('div', id='Dyn_1_3')
    Txt_2_1_3 = Data_2_5_1.find_all('a')
    for i, Txt_2_1_3 in enumerate(Txt_2_1_3):
        tmp: str = Txt_2_1_3.get_text().strip()
        tmp01: str = Txt_2_1_3.get('href')
        word4 = re.findall(r"^/p", tmp01)
        word5 = re.findall(r"tw/", tmp01)
        word6 = re.findall(r"^/var", tmp01)
        if word6:
            tmp01 = 'http://www1.chu.edu.tw' + tmp01
        if word4:
            tmp01 = 'http://www1.chu.edu.tw' + tmp01
        if word5:
            tmp01 = tmp01.rstrip('index.php?Lang=zh-tw')
        for j in enumerate(Txt_2_1_3):
            Ac_List.append(tmp)
            Ac_WEB_List.append(tmp01)
            tmp = ''
            tmp01 = ''
    # 管理學院
    Data_2_5_2 = BS_2_3.find('div', id='Dyn_1_4')
    Txt_2_1_3 = Data_2_5_2.find_all('a')
    for i, Txt_2_1_3 in enumerate(Txt_2_1_3):
        tmp: str = Txt_2_1_3.get_text().strip()
        tmp01: str = Txt_2_1_3.get('href')
        word4 = re.findall(r"^/p", tmp01)
        word5 = re.findall(r"tw/", tmp01)
        word6 = re.findall(r"^/var", tmp01)
        if word6:
            tmp01 = 'http://www1.chu.edu.tw' + tmp01
        if word4:
            tmp01 = 'http://www1.chu.edu.tw' + tmp01
        if word5:
            tmp01 = tmp01.rstrip('index.php?Lang=zh-tw')
        for j in enumerate(Txt_2_1_3):
            Ac_List.append(tmp)
            Ac_WEB_List.append(tmp01)
            tmp = ''
            tmp01 = ''
    # 建築與設計學院
    Data_2_5_3 = BS_2_3.find('div', id='Dyn_1_5')
    Txt_2_1_3 = Data_2_5_3.find_all('a')
    for i, Txt_2_1_3 in enumerate(Txt_2_1_3):
        tmp: str = Txt_2_1_3.get_text().strip()
        tmp01: str = Txt_2_1_3.get('href')
        word4 = re.findall(r"^/p", tmp01)
        word5 = re.findall(r"tw/", tmp01)
        word6 = re.findall(r"^/var", tmp01)
        if word6:
            tmp01 = 'http://www1.chu.edu.tw' + tmp01
        if word4:
            tmp01 = 'http://www1.chu.edu.tw' + tmp01
        if word5:
            tmp01 = tmp01.rstrip('index.php?Lang=zh-tw')
        for j in enumerate(Txt_2_1_3):
            Ac_List.append(tmp)
            Ac_WEB_List.append(tmp01)
            tmp = ''
            tmp01 = ''
    # 人文社會學院
    Data_2_5_4 = BS_2_3.find('div', id='Dyn_1_6')
    Txt_2_1_3 = Data_2_5_4.find_all('a')
    for i, Txt_2_1_3 in enumerate(Txt_2_1_3):
        tmp: str = Txt_2_1_3.get_text().strip()
        tmp01: str = Txt_2_1_3.get('href')
        word4 = re.findall(r"^/p", tmp01)
        word5 = re.findall(r"tw/", tmp01)
        word6 = re.findall(r"^/var", tmp01)
        if word6:
            tmp01 = 'http://www1.chu.edu.tw' + tmp01
        if word4:
            tmp01 = 'http://www1.chu.edu.tw' + tmp01
        if word5:
            tmp01 = tmp01.rstrip('index.php?Lang=zh-tw')
        for j in enumerate(Txt_2_1_3):
            Ac_List.append(tmp)
            Ac_WEB_List.append(tmp01)
            tmp = ''
            tmp01 = ''
    # 觀光學院
    Data_2_5_5 = BS_2_3.find('div', id='Dyn_1_7')
    Txt_2_1_3 = Data_2_5_5.find_all('a')
    for i, Txt_2_1_3 in enumerate(Txt_2_1_3):
        tmp: str = Txt_2_1_3.get_text().strip()
        tmp01: str = Txt_2_1_3.get('href')
        word4 = re.findall(r"^/p", tmp01)
        word5 = re.findall(r"tw/", tmp01)
        word6 = re.findall(r"^/var", tmp01)
        if word6:
            tmp01 = 'http://www1.chu.edu.tw' + tmp01
        if word4:
            tmp01 = 'http://www1.chu.edu.tw' + tmp01
        if word5:
            tmp01 = tmp01.rstrip('index.php?Lang=zh-tw')
        for j in enumerate(Txt_2_1_3):
            Ac_List.append(tmp)
            Ac_WEB_List.append(tmp01)
            tmp = ''
            tmp01 = ''
    # 創新產業學院
    Data_2_5_6 = BS_2_3.find('div', id='Dyn_1_8')
    Txt_2_1_3 = Data_2_5_6.find_all('a')
    for i, Txt_2_1_3 in enumerate(Txt_2_1_3):
        tmp: str = Txt_2_1_3.get_text().strip()
        tmp01: str = Txt_2_1_3.get('href')
        word4 = re.findall(r"^/p", tmp01)
        word5 = re.findall(r"tw/", tmp01)
        word6 = re.findall(r"^/var", tmp01)
        if word6:
            tmp01 = 'http://www1.chu.edu.tw' + tmp01
        if word4:
            tmp01 = 'http://www1.chu.edu.tw' + tmp01
        if word5:
            tmp01 = tmp01.rstrip('index.php?Lang=zh-tw')
        for j in enumerate(Txt_2_1_3):
            Ac_List.append(tmp)
            Ac_WEB_List.append(tmp01)
            tmp = ''
            tmp01 = ''
    # 通識教育中心
    Data_2_5_7 = BS_2_3.find('div', id='Dyn_1_9')
    Txt_2_1_3 = Data_2_5_7.find_all('a')
    for i, Txt_2_1_3 in enumerate(Txt_2_1_3):
        tmp: str = Txt_2_1_3.get_text().strip()
        tmp01: str = Txt_2_1_3.get('href')
        word4 = re.findall(r"^/p", tmp01)
        word5 = re.findall(r"tw/", tmp01)
        word6 = re.findall(r"^/var", tmp01)
        if word6:
            tmp01 = 'http://www1.chu.edu.tw' + tmp01
        if word4:
            tmp01 = 'http://www1.chu.edu.tw' + tmp01
        if word5:
            tmp01 = tmp01.rstrip('index.php?Lang=zh-tw')
        for j in enumerate(Txt_2_1_3):
            Ac_List.append(tmp)
            Ac_WEB_List.append(tmp01)
            tmp = ''
            tmp01 = ''
    # =================行政單位=================
    Office = requests.get(URL_offi)
    BS_2_4 = BeautifulSoup(Office.text, 'html.parser')
    # 行政單位
    Data_2_5 = BS_2_4.find('div', class_='mcont')
    Txt_2_1_4 = Data_2_5.find_all('a')
    for i, Txt_2_1_4 in enumerate(Txt_2_1_4):
        tmp: str = Txt_2_1_4.get_text().strip()
        tmp01: str = Txt_2_1_4.get('href')
        word5 = re.findall(r"^/p", tmp01)
        word6 = re.findall(r"tw/", tmp01)
        word4 = re.findall(r"^/var", tmp01)
        if word4:
            tmp01 = 'http://www1.chu.edu.tw' + tmp01
        if word5:
            tmp01 = 'http://www1.chu.edu.tw' + tmp01
        if word6:
            tmp01 = tmp01.rstrip('bin/home.php')
        for j in enumerate(Txt_2_1_4):
            Of_List.append(tmp)
            Of_WEB_List.append(tmp01)
            tmp = ''
            tmp01 = ''
    # ==========================================
    print('========開始擷取_學術與行政單位========')
    # =================學術各單位=================
    for i in range(len(Ac_List)):
        URL_ac_of = ''.join(Ac_WEB_List[i])
        unity = requests.get(URL_ac_of)
        BS_2_5 = BeautifulSoup(unity.text, 'html.parser')
        Data_2_6 = BS_2_5.find('div', class_='top_second')
        unity_bool = bool(Data_2_6)
        if unity_bool == 1:
            Txt_2_1_5 = Data_2_6.find_all('a')
            for n, Txt_2_1_5 in enumerate(Txt_2_1_5):
                tmp: str = Txt_2_1_5.get_text().strip()
                if tmp:
                    tmp01: str = Txt_2_1_5.get('href')
                    word6 = re.findall(r"^javascript", tmp01)
                    if not word6:
                        word5 = re.findall(r"^/p", tmp01)
                        word7 = re.findall(r"^/index", tmp01)
                        word4 = re.findall(r"^/var", tmp01)
                        cc += 1
                        if word4:
                            tmp01 = 'http://www1.chu.edu.tw' + tmp01
                        if word5:
                            tmp01 = URL_ac_of + tmp01
                        if word7:
                            tmp01 = ''
                        for j in enumerate(Txt_2_1_5):
                            Ac_unity_text.append(tmp)
                            Ac_unity_Web.append(tmp01)
                            tmp = ''
                            tmp01 = ''
        ac_count.append(cc)
        cc = 0
    # =================行政各單位=================
    for a in range(len(Of_List)):
        URL_ac_of = ''.join(Of_WEB_List[a])
        unity = requests.get(URL_ac_of)
        BS_2_6 = BeautifulSoup(unity.text, 'html.parser')
        Data_2_8 = BS_2_6.find('div', class_='top_second')
        unity_bool = bool(Data_2_8)
        if unity_bool == 1:
            Txt_2_1_6 = Data_2_8.find_all('a')
            for n, Txt_2_1_6 in enumerate(Txt_2_1_6):
                tmp: str = Txt_2_1_6.get_text().strip()
                if tmp:
                    tmp01: str = Txt_2_1_6.get('href')
                    word6 = re.findall(r"^javascript", tmp01)
                    if not word6:
                        word5 = re.findall(r"^/p", tmp01)
                        word7 = re.findall(r"^/index", tmp01)
                        word4 = re.findall(r"^/var", tmp01)
                        cc += 1
                        if word4:
                            tmp01 = 'http://www1.chu.edu.tw' + tmp01
                        if word5:
                            tmp01 = URL_ac_of + tmp01
                        if word7:
                            tmp01 = ''
                        for j in enumerate(Txt_2_1_6):
                            Of_unity_text.append(tmp)
                            Of_unity_Web.append(tmp01)
                            tmp = ''
                            tmp01 = ''
        of_count.append(cc)
        cc = 0
    # ==========================================
    print('========開始取得及分析_文字連結========')
    # =================學術各網頁文字搜尋==========
    web = ["main", "mainbody", "maincontent", "a", "h1", "h2", "h3"]
    # ==========學術總單位==========
    for n in range(len(Ac_WEB_List)):
        gg = Ac_WEB_List[n]
        if gg:
            try:
                u_url = requests.get(gg, headers=headers)
                u_url.encoding = 'utf8'
            except requests.exceptions.ConnectionError:
                point_ac_all_text.append('')
                point_ac_all_freq.append('')
                ac_all_sentence.append('')
            else:
                ur_tx = BeautifulSoup(u_url.text, 'html.parser')
                tmp02 = ur_tx.find_all('div', attrs={"class": web})
                if bool(tmp02) == 1:
                    for g, tmp02 in enumerate(tmp02):
                        ww: str = tmp02.get_text().strip()
                        if ww:
                            # 去除數字、英文、底線
                            ww = re.sub('[a-zA-Z0-9_]', '', ww)
                            ww = ww.replace("_", '')
                            # 替代換行，並去除其他非文字符號
                            ww = ww.replace("\n", '1')
                            ww = ww.replace(u"。", '2')
                            ww = ''.join(ch for ch, _ in itertools.groupby(ww))
                            ww = re.sub('\s+', '', ww)
                            # 以逗號區隔句子，並儲存至串列中
                            tmp11 = ww.split('1')
                            tmp11 = [x for x in tmp11 if x != '']
                            tmp12 = ','.join(str(d) for d in tmp11)
                            tmp13 = tmp12.split('2')
                            tmp13 = [x for x in tmp13 if x != '']
                            tmp14 = ','.join(str(d) for d in tmp13)
                            # 去除多餘的數字、標點符號等
                            ww = re.sub('\d', '', ww)
                            ww = re.sub('\W+', '', ww)
                            # 彙整文章、重製串列
                            tmp03 += ww
                            tmp11 = []
                    ac_all_sentence.append(tmp14)
                    # 詞語分析
                    words = [word for word in jieba.cut(tmp03) if len(word) >= 2]
                    words = [w for w in words if w not in stopWords]
                    c = Counter(words)
                    for word_freq in c.most_common(20):
                        word, freq = word_freq
                        tmp05.append(word)
                        tmp10.append(freq)
                    yy = ','.join(str(d) for d in tmp05)
                    uu = ','.join(str(d) for d in tmp10)
                    point_ac_all_text.append(yy)
                    point_ac_all_freq.append(uu)
                    tmp03 = ''
                    tmp05 = []
                    tmp10 = []
                    tmp12 = ''
                else:
                    ac_all_sentence.append('')
                    point_ac_all_text.append('')
                    point_ac_all_freq.append('')
            time.sleep(1)
        else:
            ac_all_sentence.append('')
            point_ac_all_text.append('')
            point_ac_all_freq.append('')
    # 計算次數比例
    for i, row in enumerate(point_ac_all_freq):
        if bool(point_ac_all_freq[i]) == 1:
            y = point_ac_all_freq[i].split(',')
            t = list(map(int, y))
            x = sum(t)
            for j in t:
                f = round(j / x, 3)
                tmp06.append(f)
        else:
            tmp06.append('')
        gg = ','.join(str(d) for d in tmp06)
        ratio_ac_all_freq.append(gg)
        tmp06 = []
    # ============================
    m1 = datetime.datetime.now()
    print('=========01階段完成，耗時：{}========='.format(m1 - start))
    # ==========學術分單位==========
    for n in range(len(Ac_unity_Web)):
        hh = Ac_unity_Web[n]
        if hh:
            try:
                u_url = requests.get(hh, headers=headers)
                u_url.encoding = 'utf8'
            except requests.exceptions.ConnectionError:
                point_ac_dep_text.append('')
                point_ac_dep_freq.append('')
                ac_dep_sentence.append('')
            else:
                ur_tx = BeautifulSoup(u_url.text, 'html.parser')
                tmp02 = ur_tx.find_all('div', attrs={"class": web})
                if bool(tmp02) == 1:
                    for g, tmp02 in enumerate(tmp02):
                        ww: str = tmp02.get_text().strip()
                        if ww:
                            # 去除數字、英文、底線
                            ww = re.sub('[a-zA-Z0-9_]', '', ww)
                            ww = ww.replace("_", '')
                            # 替代換行，並去除其他非文字符號
                            ww = ww.replace("\n", '1')
                            ww = ww.replace(u"。", '2')
                            ww = ''.join(ch for ch, _ in itertools.groupby(ww))
                            ww = re.sub('\s+', '', ww)
                            # 以逗號區隔句子，並儲存至串列中
                            tmp11 = ww.split('1')
                            tmp11 = [x for x in tmp11 if x != '']
                            tmp12 = ','.join(str(d) for d in tmp11)
                            tmp13 = tmp12.split('2')
                            tmp13 = [x for x in tmp13 if x != '']
                            tmp14 = ','.join(str(d) for d in tmp13)
                            # 去除多餘的數字、標點符號等
                            ww = re.sub('\d', '', ww)
                            ww = re.sub('\W+', '', ww)
                            # 彙整文章、重製串列
                            tmp03 += ww
                            tmp11 = []
                    ac_dep_sentence.append(tmp14)
                    # 詞語分析
                    words = [word for word in jieba.cut(tmp03) if len(word) >= 2]
                    words = [w for w in words if w not in stopWords]
                    c = Counter(words)
                    for word_freq in c.most_common(20):
                        word, freq = word_freq
                        tmp05.append(word)
                        tmp10.append(freq)
                    yy = ','.join(str(d) for d in tmp05)
                    uu = ','.join(str(d) for d in tmp10)
                    point_ac_dep_text.append(yy)
                    point_ac_dep_freq.append(uu)
                    tmp03 = ''
                    tmp05 = []
                    tmp10 = []
                    tmp12 = ''
                else:
                    ac_dep_sentence.append('')
                    point_ac_dep_text.append('')
                    point_ac_dep_freq.append('')
            time.sleep(1)
        else:
            ac_dep_sentence.append('')
            point_ac_dep_text.append('')
            point_ac_dep_freq.append('')
    # 計算次數比例
    for i, row in enumerate(point_ac_dep_freq):
        if bool(point_ac_dep_freq[i]) == 1:
            y = point_ac_dep_freq[i].split(',')
            t = list(map(int, y))
            x = sum(t)
            for j in t:
                f = round(j / x, 3)
                tmp06.append(f)
        else:
            tmp06.append('')
        gg = ','.join(str(d) for d in tmp06)
        ratio_ac_dep_freq.append(gg)
        tmp06 = []
    # ============================
    m2 = datetime.datetime.now()
    print('=========02階段完成，耗時：{}========='.format(m2 - m1))
    # ==========行政總單位==========
    for n in range(len(Of_WEB_List)):
        ee = Of_WEB_List[n]
        if ee:
            try:
                u_url = requests.get(ee, headers=headers)
                u_url.encoding = 'utf8'
            except requests.exceptions.ConnectionError:
                point_of_all_text.append('')
                point_of_all_freq.append('')
                of_all_sentence.append('')
            else:
                ur_tx = BeautifulSoup(u_url.text, 'html.parser')
                tmp02 = ur_tx.find_all('div', attrs={"class": web})
                if bool(tmp02) == 1:
                    for g, tmp02 in enumerate(tmp02):
                        ww: str = tmp02.get_text().strip()
                        if ww:
                            # 去除數字、英文、底線
                            ww = re.sub('[a-zA-Z0-9_]', '', ww)
                            ww = ww.replace("_", '')
                            # 替代換行，並去除其他非文字符號
                            ww = ww.replace("\n", '1')
                            ww = ww.replace(u"。", '2')
                            ww = ''.join(ch for ch, _ in itertools.groupby(ww))
                            ww = re.sub('\s+', '', ww)
                            # 以逗號區隔句子，並儲存至串列中
                            tmp11 = ww.split('1')
                            tmp11 = [x for x in tmp11 if x != '']
                            tmp12 = ','.join(str(d) for d in tmp11)
                            tmp13 = tmp12.split('2')
                            tmp13 = [x for x in tmp13 if x != '']
                            tmp14 = ','.join(str(d) for d in tmp13)
                            # 去除多餘的數字、標點符號等
                            ww = re.sub('\d', '', ww)
                            ww = re.sub('\W+', '', ww)
                            # 彙整文章、重製串列
                            tmp03 += ww
                            tmp11 = []
                    of_all_sentence.append(tmp14)
                    # 詞語分析
                    words = [word for word in jieba.cut(tmp03) if len(word) >= 2]
                    words = [w for w in words if w not in stopWords]
                    c = Counter(words)
                    for word_freq in c.most_common(20):
                        word, freq = word_freq
                        tmp05.append(word)
                        tmp10.append(freq)
                    yy = ','.join(str(d) for d in tmp05)
                    uu = ','.join(str(d) for d in tmp10)
                    point_of_all_text.append(yy)
                    point_of_all_freq.append(uu)
                    tmp03 = ''
                    tmp05 = []
                    tmp10 = []
                    tmp12 = ''
                else:
                    of_all_sentence.append('')
                    point_of_all_text.append('')
                    point_of_all_freq.append('')
            time.sleep(1)
        else:
            of_all_sentence.append('')
            point_of_all_text.append('')
            point_of_all_freq.append('')
    # 計算次數比例
    for i, row in enumerate(point_of_all_freq):
        if bool(point_of_all_freq[i]) == 1:
            y = point_of_all_freq[i].split(',')
            t = list(map(int, y))
            x = sum(t)
            for j in t:
                f = round(j / x, 3)
                tmp06.append(f)
        else:
            tmp06.append('')
        gg = ','.join(str(d) for d in tmp06)
        ratio_of_all_freq.append(gg)
        tmp06 = []
    # ============================
    m3 = datetime.datetime.now()
    print('=========03階段完成，耗時：{}========='.format(m3 - m2))
    # ==========行政分單位==========
    for n in range(len(Of_unity_Web)):
        ff = Of_unity_Web[n]
        if ff:
            try:
                u_url = requests.get(ff, headers=headers)
                u_url.encoding = 'utf8'
            except requests.exceptions.ConnectionError:
                point_of_dep_text.append('')
                point_of_dep_freq.append('')
                of_dep_sentence.append('')
            else:
                ur_tx = BeautifulSoup(u_url.text, 'html.parser')
                tmp02 = ur_tx.find_all('div', attrs={"class": web})
                if bool(tmp02) == 1:
                    for g, tmp02 in enumerate(tmp02):
                        ww: str = tmp02.get_text().strip()
                        if ww:
                            # 去除數字、英文、底線
                            ww = re.sub('[a-zA-Z0-9_]', '', ww)
                            ww = ww.replace("_", '')
                            # 替代換行，並去除其他非文字符號
                            ww = ww.replace("\n", '1')
                            ww = ww.replace(u"。", '2')
                            ww = ''.join(ch for ch, _ in itertools.groupby(ww))
                            ww = re.sub('\s+', '', ww)
                            # 以逗號區隔句子，並儲存至串列中
                            tmp11 = ww.split('1')
                            tmp11 = [x for x in tmp11 if x != '']
                            tmp12 = ','.join(str(d) for d in tmp11)
                            tmp13 = tmp12.split('2')
                            tmp13 = [x for x in tmp13 if x != '']
                            tmp14 = ','.join(str(d) for d in tmp13)
                            # 去除多餘的數字、標點符號等
                            ww = re.sub('\d', '', ww)
                            ww = re.sub('\W+', '', ww)
                            # 彙整文章、重製串列
                            tmp03 += ww
                            tmp11 = []
                    of_dep_sentence.append(tmp14)
                    # 詞語分析
                    words = [word for word in jieba.cut(tmp03) if len(word) >= 2]
                    words = [w for w in words if w not in stopWords]
                    c = Counter(words)
                    for word_freq in c.most_common(20):
                        word, freq = word_freq
                        tmp05.append(word)
                        tmp10.append(freq)
                    yy = ','.join(str(d) for d in tmp05)
                    uu = ','.join(str(d) for d in tmp10)
                    point_of_dep_text.append(yy)
                    point_of_dep_freq.append(uu)
                    tmp03 = ''
                    tmp05 = []
                    tmp10 = []
                    tmp12 = ''
                else:
                    of_dep_sentence.append('')
                    point_of_dep_text.append('')
                    point_of_dep_freq.append('')
            time.sleep(1)
        else:
            of_dep_sentence.append('')
            point_of_dep_text.append('')
            point_of_dep_freq.append('')
    # 計算次數比例
    for i, row in enumerate(point_of_dep_freq):
        if bool(point_of_dep_freq[i]) == 1:
            y = point_of_dep_freq[i].split(',')
            t = list(map(int, y))
            x = sum(t)
            for j in t:
                f = round(j / x, 3)
                tmp06.append(f)
        else:
            tmp06.append('')
        gg = ','.join(str(d) for d in tmp06)
        ratio_of_dep_freq.append(gg)
        tmp06 = []
    # ============================
    m4 = datetime.datetime.now()
    print('=========04階段完成，耗時：{}========='.format(m4 - m3))
    # ============================
    a = len(point_ac_all_text)
    b = len(point_ac_dep_text)
    c = len(point_of_all_text)
    d = len(point_of_dep_text)
    print(len(ac_all_sentence), ',', len(ac_dep_sentence), ',', len(of_all_sentence), ',', len(of_dep_sentence))
    print('01階段共：{} 個, 02階段共：{} 個, 03階段共：{} 個, 04階段共：{} 個, 總數：{} 個'.format(a, b, c, d, a + b + c + d))
    # =================Excel==================
    # =================儲存資料=================
    x_01 = bool(wb.worksheets[0])
    x_02 = bool(wb.worksheets[1])
    x_03 = bool(wb.worksheets[2])
    # ==========================================================================
    # 輸出至第一資料表
    if x_01 == 1:
        wb.remove(wb.worksheets[0])
        ws = wb.create_sheet('data_01', 0)
        ws['A1'] = '序號'
        ws['B1'] = '學術單位名稱'
        ws['C1'] = '子名稱'
        ws['D1'] = '網頁連結'
        ws['E1'] = '關鍵字_詞彙'
        ws['F1'] = '關鍵字_次數'
        ws['G1'] = '關鍵字_摘句'
        for i in range(len(Ac_List)):
            ws.append(['A_1_' + str(i + 1), Ac_List[i], '學術', Ac_WEB_List[i],
                       point_ac_all_text[i], ratio_ac_all_freq[i], ac_all_sentence[i]])
        for j in range(len(Ac_unity_text)):
            ws.append(['A_2_' + str(j + 1), '', Ac_unity_text[j], Ac_unity_Web[j],
                       point_ac_dep_text[j], ratio_ac_dep_freq[j], ac_dep_sentence[j]])
        tt = 0
        for k in range(len(ac_count)):
            gg = ac_count[k]
            tmp07 = Ac_List[k]
            if gg != 0:
                for m in range(gg):
                    ws.cell(len(Ac_List) + 2 + m + tt, 2, tmp07)
            tt += gg
    # ==========================================================================
    # 輸出至第二資料表
    if x_02 == 1:
        wb.remove(wb.worksheets[1])
        ws = wb.create_sheet('data_02', 1)
        ws['A1'] = '序號'
        ws['B1'] = '行整單位名稱'
        ws['C1'] = '子名稱'
        ws['D1'] = '網頁連結'
        ws['E1'] = '關鍵字_詞彙'
        ws['F1'] = '關鍵字_次數'
        ws['G1'] = '關鍵字_摘句'
        for i in range(len(Of_List)):
            ws.append(['B_1_' + str(i + 1), Of_List[i], '行政', Of_WEB_List[i],
                       point_of_all_text[i], ratio_of_all_freq[i], of_all_sentence[i]])
        for j in range(len(Of_unity_text)):
            ws.append(['B_2_' + str(j + 1), '', Of_unity_text[j], Of_unity_Web[j],
                       point_of_dep_text[j], ratio_of_dep_freq[j], of_dep_sentence[j]])
        tt = 0
        for k in range(len(of_count)):
            gg = of_count[k]
            tmp07 = Of_List[k]
            if gg != 0:
                for m in range(gg):
                    ws.cell(len(Of_List) + 2 + m + tt, 2, tmp07)
            tt += gg
    # ==========================================================================
    # 輸出至第三資料表
    if x_03 == 1:
        wb.remove(wb.worksheets[2])
        ws = wb.create_sheet('data_03', 2)
        ws['A1'] = '序號'
        ws['B1'] = '單位名稱'
        ws['C1'] = '網頁連結'
        for i in range(len(Ac_List)):
            ws.append(['A_1_' + str(i + 1), Ac_List[i], '學術', Ac_WEB_List[i]])
        for j in range(len(Of_List)):
            ws.append(['B_1_' + str(j + 1), Of_List[j], '行政', Of_WEB_List[j]])
    # ==========================================================================
    wb.save('resource/cra_data.xlsx')
    print('========試算表_儲存資料完成===========')
    # ==========================================================================
    # 結束 - 執行時間
    end = datetime.datetime.now()
    total_time = end - start
    # 輸出時間結果
    print("爬蟲總執行時間：{}".format(total_time))
    # ==========================================================================


# 中華大學網站 資料庫副程式
def sql():
    # ==========================================================================
    # 起始 - 執行時間
    print('========資料庫 開始執行==============')
    start = datetime.datetime.now()
    # ==========================================================================
    # SQL-lite - 攜帶式資料庫
    db = sqlite3.connect('resource/python_E_Final.db')
    cursor = db.cursor()
    # ========================================
    print('========資料庫 刪除現有資料===========')
    # ========================================
    # 刪除資料庫所有資料
    cursor.execute("DELETE FROM keys;")
    db.commit()
    cursor.execute("DELETE FROM sqlite_sequence WHERE name = 'keys';")
    db.commit()
    # ========================================
    s1 = datetime.datetime.now()
    print('=========01階段完成，耗時：{}========='.format(s1 - start))
    # ========================================
    # 執行Excel檔
    wb2 = load_workbook('resource/cra_data.xlsx')
    # ========================================
    # 資料庫 keys資料表
    print('========資料庫 學術 加入keys資料表====')
    # 學術單位
    tmp01 = []  # 關鍵詞
    tmp02 = []  # 隸屬之一
    tmp08 = []  # 隸屬之二
    tmp03 = []  # 網站
    tmp04 = []  # 關鍵詞的比例
    tmp05 = []  # 暫存容器
    tmp06 = ''  # 暫存容器
    tmp07 = []  # 關鍵詞摘句
    pp = 2
    # 匯出Excel檔案
    w_a = wb2.worksheets[0]
    if bool(w_a) == 1:
        for i in w_a['A']:
            if i.value != 'None':
                tmp01.append(w_a.cell(pp, 5).value)
                tmp02.append(w_a.cell(pp, 2).value)
                tmp08.append(w_a.cell(pp, 3).value)
                tmp03.append(w_a.cell(pp, 4).value)
                tmp04.append(w_a.cell(pp, 6).value)
                tmp07.append(w_a.cell(pp, 7).value)
            else:
                tmp01.append('')
                tmp02.append('')
                tmp03.append('')
                tmp04.append('')
                tmp07.append('')
            pp += 1
    # 輸出至資料庫
    for i, row1 in enumerate(tmp01):
        if bool(tmp01[i]) == 1:
            keys = tmp01[i].split(',')
            ratio = tmp04[i].split(',')
            sentence = ''.join(tmp07[i])
            for j, row2 in enumerate(keys):
                # 輸入至暫存串列
                tmp05.append(keys[j])
                tmp05.append(tmp02[i])
                tmp05.append(tmp08[i])
                tmp05.append(tmp03[i])
                tmp05.append(sentence)
                tmp05.append(ratio[j])
                # ========================================
                rows = [tmp05[0], tmp05[1], tmp05[2], tmp05[3], tmp05[4], tmp05[5], 0]
                cursor.execute("INSERT INTO keys "
                               "(keywords, depart, subpart, website, sentence, ratio, second_ratio) "
                               "values(?,?,?,?,?,?,?);", rows)
                db.commit()
                # ========================================
                tmp05 = []
    # ========================================
    s2 = datetime.datetime.now()
    print('=========02階段完成，耗時：{}========='.format(s2 - s1))
    # ========================================
    print('========資料庫 行政 加入keys資料表====')
    # 行政單位
    tmp01 = []  # 關鍵詞
    tmp02 = []  # 隸屬之一
    tmp08 = []  # 隸屬之二
    tmp03 = []  # 網站
    tmp04 = []  # 關鍵詞的比例
    tmp05 = []  # 暫存容器
    tmp06 = ''  # 暫存容器
    tmp07 = []  # 關鍵詞摘句
    qq = 2
    # 匯出Excel檔案
    w_a = wb2.worksheets[1]
    if bool(w_a) == 1:
        for i in w_a['A']:
            if i.value != 'None':
                tmp01.append(w_a.cell(qq, 5).value)
                tmp02.append(w_a.cell(qq, 2).value)
                tmp08.append(w_a.cell(qq, 3).value)
                tmp03.append(w_a.cell(qq, 4).value)
                tmp04.append(w_a.cell(qq, 6).value)
                tmp07.append(w_a.cell(qq, 7).value)
            else:
                tmp01.append('')
                tmp02.append('')
                tmp03.append('')
                tmp04.append('')
                tmp07.append('')
            qq += 1
    # 輸出至資料庫
    for i, row1 in enumerate(tmp01):
        if bool(tmp01[i]) == 1:
            keys = tmp01[i].split(',')
            ratio = tmp04[i].split(',')
            sentence = ''.join(tmp07[i])
            for j, row2 in enumerate(keys):
                # 輸入至暫存串列
                tmp05.append(keys[j])
                tmp05.append(tmp02[i])
                tmp05.append(tmp08[i])
                tmp05.append(tmp03[i])
                tmp05.append(sentence)
                tmp05.append(ratio[j])
                # ========================================
                rows = [tmp05[0], tmp05[1], tmp05[2], tmp05[3], tmp05[4], tmp05[5], 0]
                cursor.execute("INSERT INTO keys "
                               "(keywords, depart, subpart, website, sentence, ratio, second_ratio) "
                               "values(?,?,?,?,?,?,?);", rows)
                db.commit()
                # ========================================
                tmp05 = []
    # ========================================
    s3 = datetime.datetime.now()
    print('=========03階段完成，耗時：{}========='.format(s3 - s2))
    # ========================================
    print('========資料庫 單位 加入keys資料表===')
    # 資料庫 other_客製插入
    other_list_key = ['首頁 - 中華大學', '在校生', '在校生', '在校生', '在校生', '在校生', '在校生']
    other_list_dep = ['首頁 - 中華大學', '數位學習平台', 'Web_mail信箱', '課程查詢系統', '註冊繳費e化', '聯合服務中心', '學生資訊系統']
    other_list_sub = ['其他單位', '其他單位', '其他單位', '其他單位', '其他單位', '其他單位', '其他單位']
    other_list_web = ['https://www1.chu.edu.tw/', 'https://elearn.chu.edu.tw/', 'https://mail.chu.edu.tw/',
                      'http://course.chu.edu.tw/', 'https://student.chu.edu.tw/non_register/non-register.htm',
                      'http://ga.chu.edu.tw/p/412-1051-2197.php?Lang=zh-tw', 'https://student2.chu.edu.tw/']
    other_list_sen = ['首頁 - 中華大學', '數位學習平台', 'Web_mail信箱', '課程查詢系統', '註冊繳費e化', '聯合服務中心', '學生資訊系統']
    other_list_rai = ['10', '10', '10', '10', '10', '10', '10']
    # 學術總單位
    for q in range(len(other_list_key)):
        rows = [other_list_key[q], other_list_dep[q], other_list_sub[q], other_list_web[q],
                other_list_sen[q], other_list_rai[q], 0]
        cursor.execute("INSERT INTO keys (keywords, depart, subpart, website, sentence, ratio, second_ratio) "
                       "values(?,?,?,?,?,?,?);", rows)
        db.commit()
    # ========================================
    s4 = datetime.datetime.now()
    print('=========04階段完成，耗時：{}========='.format(s4 - s3))
    # ========================================
    # 關閉資料庫
    print('========資料庫 關閉=================')
    cursor.close()
    db.close()
    # ==========================================================================
    # 結束 - 執行時間
    end = datetime.datetime.now()
    total_time = end - start
    # 輸出時間結果
    print("資料庫總執行時間：{}".format(total_time))
    # ==========================================================================


# 中華大學網站 資料庫二階整理
def second_db():
    # ==========================================================================
    # 起始 - 執行時間
    print('========二階資料整理 開始執行==============')
    start = datetime.datetime.now()
    # ==========================================================================
    # SQL-lite - 攜帶式資料庫
    db = sqlite3.connect('resource/python_E_Final.db')
    cursor = db.cursor()
    # ==========================================================================
    # 取得全部不同的關鍵詞
    cursor.execute("SELECT keywords FROM keys GROUP by keywords;")
    keywords_list = [i[0] for i in cursor.fetchall()]
    # ==========================================================================
    print('====資料彙整完畢，開始進行計算比例====')
    # 取得各組關鍵詞的 ID、關鍵詞、比例
    for m in range(len(keywords_list)):
        d = {}
        keywords_dict = {}
        total = 0
        db_str_01 = "SELECT id, ratio FROM keys WHERE keywords = '{0}';".format(keywords_list[m])
        cursor.execute(db_str_01)
        rows = cursor.fetchall()
        for row in rows:
            keywords_dict.update({row[0]: row[1]})
        # ========================================
        # 計算當前關鍵詞的總比例
        for tmp02 in keywords_dict.values():
            total += tmp02
        # 計算當前關鍵詞的個別比例並放入資料庫
        for key, value in keywords_dict.items():
            tmp04 = value / total
            tmp04 = round(tmp04, 3)
            d.update({key: tmp04})
        # ========================================
        for d_key, d_value in d.items():
            db_str_02 = "UPDATE keys SET second_ratio = {0} WHERE id = {1};".format(d_value, d_key)
            cursor.execute(db_str_02)
            db.commit()
    # ========================================
    # 關閉資料庫
    print('========資料庫 關閉=================')
    cursor.close()
    db.close()
    # ==========================================================================
    # 結束 - 執行時間
    end = datetime.datetime.now()
    total_time = end - start
    # 輸出時間結果
    print("資料庫總執行時間：{}".format(total_time))
    # ==========================================================================


# 中華大學網站 資料庫摘句整理
def sentence_ratio():
    # ==========================================================================
    # 起始 - 執行時間
    print('========資料庫摘句整理 開始執行==============')
    start = datetime.datetime.now()
    # ==========================================================================
    # SQL-lite - 攜帶式資料庫
    db = sqlite3.connect('resource/python_E_Final.db')
    cursor = db.cursor()
    # ==========================================================================
    # 取得網頁數量
    web_text = []
    web = []
    cursor.execute("SELECT website, count(keywords) FROM keys GROUP by depart, subpart, website ORDER by id;")
    rowe = cursor.fetchall()
    for r in rowe:
        web_text.append(r[0])
        web.append(r[1])
    # ==========================================================================
    print('====資料彙整完畢，開始進行計算正理====')
    # 取得同網站的關鍵詞
    for i, row in enumerate(web):
        # 定義
        id_list = []
        key_list = []
        sen_list = []
        # 取得資料
        db_str = "SELECT id, keywords, sentence FROM keys WHERE website is '{0}' ORDER by id;".format(web_text[i])
        print('第', i + 1, '項。', db_str)
        cursor.execute(db_str)
        rows = cursor.fetchall()
        for r in rows:
            id_list.append(r[0])
            key_list.append(r[1])
            sen_list.append(r[2])
        # 搜尋關鍵字所在句子，並濃縮句子
        sen_list = sen_list[0].split(',')
        for j, row1 in enumerate(id_list):
            re_dict = {sen_list[i]: 0 for i in [i for i, x in enumerate(sen_list) if x.find(key_list[j]) != -1]}
            fa_dict = {sen_list[i]: 0 for i in [i for i, x in enumerate(sen_list) if x.find(key_list[j]) == -1]}
            re_dict.update(fa_dict)
            # 尋找同網頁關鍵詞是否有出現
            for key_count, row2 in enumerate(id_list):
                for key, val in re_dict.items():
                    if key_list[key_count] in key:
                        re_dict[key] += 1
            sort_dict = sorted(re_dict.items(), key=lambda x: x[1], reverse=True)
            sort_list = [sort_dict[i][0] for i, x in enumerate(sort_dict)]
            # 擷取句子
            word = ''
            for n in range(len(sort_list)):
                pp = sort_list[n]
                pp = '{:.30}'.format(pp)
                word += pp + "。"
            word += "......"
            msg = '{:.100}'.format(word)
            # 更新資料庫
            ff = (msg, id_list[j])
            cursor.execute("UPDATE keys SET sentence = ? WHERE id = ?;", ff)
            print(id_list[j])
            db.commit()
            time.sleep(0.05)
    # ========================================
    # 關閉資料庫
    print('========資料庫 關閉=================')
    cursor.close()
    db.close()
    # ==========================================================================
    # 結束 - 執行時間
    end = datetime.datetime.now()
    total_time = end - start
    # 輸出時間結果
    print("資料庫總執行時間：{}".format(total_time))
    # ==========================================================================


# ==========================================================================
# 執行主程式
if __name__ == '__main__':
    user_input = 0
    while True:
        user_input = int(input("1: 完整執行，2:爬蟲函式，3: 資料庫函式，4: 二階整理函式，5: 摘句計算，6:結束程式\n"
                               "請輸入對應代碼開始程序>> "))
        if user_input == 1:
            print("執行總程式，預計耗時3.5小時")
            cra()
            sql()
            second_db()
            sentence_ratio()
            print("完成")
        if user_input == 2:
            print("執行爬蟲函式，預計耗時30分鐘")
            cra()
            print("完成")
        if user_input == 3:
            print("執行資料庫函式，預計耗時40分鐘")
            sql()
            print("完成")
        if user_input == 4:
            print("執行二階資料庫整理函式，預計耗時1.5小時")
            second_db()
            print("完成")
        if user_input == 5:
            print("執行二階資料庫整理函式，預計耗時1小時")
            sentence_ratio()
            print("完成")
        if user_input == 6:
            print("使用者結束程式")
            sys.exit()
# ==========================================================================
